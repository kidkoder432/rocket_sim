import openpyxl
from pid import PID
import matplotlib.pyplot as plt

eo, oo = [], []

wb = openpyxl.load_workbook("rocket_data.xlsx", read_only=True)

print(wb.sheetnames)


def scale(x, in_min, in_max, out_min, out_max):
    return (x - in_min) * (out_max - out_min) / (in_max - in_min) + out_min


sheet = wb["Sheet2"]

allErors = 0

lastTime = 2.65
for row in sheet.iter_rows(values_only=True):
    xin, xout = float(row[-1]), float(row[1])
    yin, yout = float(row[2]), float(row[3])
    t = float(row[0])
    dt = t - lastTime

    if xout == 95 and yout == 85:
        continue

    pid = PID(Kp=1.2, Ki=.1, Kd=.1, N=50, dt=dt)
    pid.limits = (-5, 5)
    pid.setpoint = 0
    expected_xout = pid(xin, dt)

    xout = scale(xout, 83, 107, -5, 5)

    allErors += abs(xout - expected_xout)

    print(
        f"X: Input: {xin}, Expected Output: {expected_xout}, Output (scaled): {xout}, Error: {xout - expected_xout}, Time: {t}"
    )
    lastTime = t

    eo.append(expected_xout)
    oo.append(xout)

print("Average error: ", allErors / sheet.max_row)

xs = range(len(eo))
plt.plot(xs, eo, oo)
plt.legend(["Expected", "Measured"])

plt.show()
